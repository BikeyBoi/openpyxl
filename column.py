import openpyxl

# Load the Excel file into a workbook
wb = openpyxl.load_workbook(r'<excel sheet path>') 

# Select the active sheet
ws = wb.active

while True:
    
    id_num = input('ID (or type "exit" or "q" to end): ')

    
    if id_num.lower() in ('exit', 'q'):
        break

    # Convert ID to row number
    row_id = None

    # Search for the row based on the entered ID in column "C3" for example
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=3, max_col=3):
        if str(row[0].value) == id_num:
            row_id = row[0].row
            break

    if row_id is not None:
        # Get the entire row for the specified ID
        result = ws[row_id]

        # Get column names from the header row (assuming header is in row 3)
        header_row = ws[3]
        column_names = [cell.value for cell in header_row]

        # Print the entire row
        for column_name, cell in zip(column_names, result):
            print(f' {column_name}: {cell.value}')
    else:
        print(f" '{id_num}' not found.")


print("Exiting the program.")
