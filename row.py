import openpyxl

# Load the Excel file into a workbook
wb = openpyxl.load_workbook(r'excel sheet path')

# Select the active sheet
ws = wb.active

while True:
    
    try:
        station_id = input("Enter ID (type 'exit' or 'q' to end): ")
        # Check if the user wants to exit
        if station_id.lower() == 'q' or station_id.lower() == 'exit':
            print("Exiting the program.")
            break
        station_id = int(station_id)  # Convert to integer if possible
    except ValueError:
        print("Invalid input. Please enter a valid integer or type 'e' or 'quit' to end.")
        continue

    # Search for the column index based on the entered ID in row 5 for example
    column_index = None
    for cell in ws[5]:
        if station_id == cell.value:
            column_index = cell.column
            break  

    if column_index is not None:
        # Get the corresponding values
        station_values = []
        for row in range(3, ws.max_row + 1):
            value = ws.cell(row=row, column=column_index).value
            header = ws.cell(row=row, column=1).value  # Assuming the first cell in column A is the header
            if value and header:
                station_values.append(f"{header}: {value}")

        
        for info in station_values:
            print(info)
    else:
        print(f"{station_id} not found in row 5.")
